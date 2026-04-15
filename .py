from openpyxl import load_workbook

archivo = r"C:\Users\pe.intern\Desktop\MB-BMM-65A60QUV(02).xlsx"

wb = load_workbook(archivo)
ws = wb.active

headers = [cell.value for cell in ws[1]]

col_level = headers.index("LEVEL") + 1
col_sort = headers.index("SORTSTRNG") + 1 if "SORTSTRNG" in headers else None

if col_sort is None:
    col_sort = len(headers) + 1
    ws.cell(row=1, column=col_sort, value="SORTSTRNG")

max_row = ws.max_row
max_col = ws.max_column

filas_a_eliminar = set()

print("Procesando bloques...\n")

i = 2

while i <= max_row:
    level_actual = ws.cell(row=i, column=col_level).value
    inicio = i

    while i <= max_row and ws.cell(row=i, column=col_level).value == level_actual:
        i += 1

    fin = i
    tamaño = fin - inicio

    # 👉 PRINT PRINCIPAL
    print(f"LEVEL {level_actual} | tamaño {tamaño}")

    if tamaño == 6:

        # Buscar padre
        fila_padre = None
        for j in range(inicio - 1, 1, -1):
            if ws.cell(row=j, column=col_level).value < level_actual:
                fila_padre = j
                break

        if fila_padre is None:
            print("→ No se encontró padre, se omite\n")
            continue

        # Marcar AI
        for fila in range(inicio, fin):
            ws.cell(row=fila, column=col_sort).value = "AI"

        # Copiar padre a última fila
        fila_destino = fin - 1

        for col in range(1, max_col + 1):
            origen = ws.cell(row=fila_padre, column=col)
            destino = ws.cell(row=fila_destino, column=col)

            destino.value = origen.value

            if origen.has_style:
                destino._style = origen._style

        ws.cell(row=fila_destino, column=col_sort).value = "AI"

        # Marcar para eliminar
        filas_a_eliminar.add(fila_padre)

        # 👉 PRINT DETALLADO
        print(f"→ Bloque procesado ({inicio}-{fin-1}) | Padre en fila {fila_padre}")

# Eliminar filas
for fila in sorted(filas_a_eliminar, reverse=True):
    print(f"Eliminando fila padre: {fila}")
    ws.delete_rows(fila)

# Guardar
output = archivo.replace(".xlsx", "_procesado.xlsx")
wb.save(output)

print(f"\nArchivo guardado en: {output}")