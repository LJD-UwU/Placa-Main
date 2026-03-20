from backend.config.credenciales_loader import cargar_credenciales, guardar_credenciales
from backend.modules.procesar_mainboard_P2 import procesar_material_desde_mainboard
from backend.modules.extract_mainboard import extract_descripcion_numbers
from backend.modules.procesar_motherboard_P1 import procesar_number
from backend.utils.clean_excel import limpiar_excel_mainboard
from backend.config.sap_login import abrir_sap_y_login
from tkinter import ttk, filedialog, messagebox, scrolledtext
from PIL import Image, ImageTk
import os, re, time, sys,subprocess
from openpyxl import load_workbook
import tkinter as tk
import pandas as pd
from datetime import datetime
from backend.Helpers.helper import cargar_archivos_procesados,guardar_archivo_procesado
from backend.modules.cs11 import ejecutar_cs11
from backend.utils.txt_to_xlsx import(
    exportar_bom_a_xls,
    convertir_xls_a_xlsx,
    BASE_BOM_FOLDER,
    MODEL_FILES_FOLDER,
    MAINBOARD_1_FILES_FOLDER,
    MAINBOARD_2_FILES_FOLDER,
)
from backend.modules.procesar_motherboard_P1 import actualizar_excel_mainboard_1
from backend.modules.procesar_mainboard_P2 import actualizar_excel_mainboard_2
from backend.config.sap_config import (
    DESCRIPCIONES,
    FILTRO,
)

class SAPApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MBAutomator")
        self.root.geometry("570x480") 
        self.root.resizable(False, False)

        #! Icono
        try:
            img = Image.open("backend/IMG/logo.png") 
            img = img.resize((64, 64))  
            icon = ImageTk.PhotoImage(img)
            self.root.iconphoto(True, icon)
        except Exception as e:
            print(f"No se pudo cargar el icono: {e}")

        #! Animación y tiempo
        self.animando = False
        self.anim_dots = 0
        self.start_time = None

        #! Estilo
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel", font=("Segoe UI", 14, "bold"))
        style.configure("TProgressbar", thickness=12)

        #!  Títulos 
        ttk.Label(root, text="Automatización SAP", style="Title.TLabel").pack(pady=(8, 0))
        ttk.Label(root, text="Procesamiento automático para el BOM", foreground="gray").pack(pady=(0, 8))

        #!  Frame principal 
        main = ttk.Frame(root, padding=6)
        main.pack(fill="both", expand=True)

        #!  Selección de Excel 
        fila_file = ttk.Frame(main)
        fila_file.pack(fill="x", pady=4)
        self.excel_path = tk.StringVar()
        ttk.Entry(fila_file, textvariable=self.excel_path).pack(side="left", fill="x", expand=True)
        ttk.Button(fila_file, text="📂", width=3, command=self.seleccionar_excel).pack(side="left", padx=4)

        #!  Barra de progreso 
        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.pack(fill="x", pady=6)

        #!  Botones principales 
        fila_btn = ttk.Frame(main)
        fila_btn.pack(pady=6)

        self.btn_mainboard = ttk.Button(fila_btn, text="🖥 Procesar Motherboard", command=self.abrir_app_mainboard, state="disabled")
        self.btn_mainboard.pack(side="left", padx=4)

        self.btn_procesar = ttk.Button(fila_btn, text="▶ Procesar 1TE", command=self.iniciar, state="disabled")
        self.btn_procesar.pack(side="left", padx=4)

        self.btn_limpiar = ttk.Button(fila_btn, text="🧹 Limpiar", command=self.limpiar_datos)
        self.btn_limpiar.pack(side="left", padx=4)

        self.btn_open = ttk.Button(fila_btn, text="📁 Resultados", command=self.abrir_resultados, state="disabled")
        self.btn_open.pack(side="left", padx=4)

        self.btn_credenciales = ttk.Button(fila_btn, text="🔐 Login SAP", command=self.abrir_credenciales)
        self.btn_credenciales.pack(side="left", padx=4)

        #!  Consola 
        frame_log = ttk.LabelFrame(main, text="CONSOLA")
        frame_log.pack(fill="both", expand=True, pady=(6, 0))
        self.log = scrolledtext.ScrolledText(frame_log, height=10, font=("Consolas", 10))
        self.log.pack(fill="both", expand=True, padx=5, pady=5)
        self.log.config(state="disabled")
        self.log.tag_config("INFO", foreground="blue")
        self.log.tag_config("OK", foreground="green", font=("Consolas", 10, "bold"))
        self.log.tag_config("ERROR", foreground="red", font=("Consolas", 10, "bold"))
        self.log.tag_config("WARNING", foreground="orange", font=("Consolas", 10, "italic"))

        #!  Estado con porcentaje 
        self.status = tk.StringVar(value="Estado: Listo")
        self.progress_label = ttk.Label(root, textvariable=self.status, anchor="w")
        self.progress_label.pack(fill="x", side="bottom", padx=6, pady=4)

        #!  Datos internos 
        self.modelos = []
        self.idx = 0
        self.session = None
        self.df_todos = pd.DataFrame(columns=["Modelo", "Planta", "Number", "Descripcion"])
        self.materiales_procesados_ok = []

        #!  Vigilar cambios en Excel 
        self.excel_path.trace_add("write", lambda *args: self.verificar_habilitar_botones())
        self.verificar_habilitar_botones()

        #!  Tooltips de botones 
        self._crear_tooltips()
        
    def _crear_tooltips(self):
            tooltips = {
                self.btn_mainboard: "Procesar archivo desde las motherboard",
                self.btn_procesar: "Procesar archivo con 1TE",
                self.btn_limpiar: "Limpiar la consola y archivos finales",
                self.btn_open: "Abrir carpeta de los archivo",
                self.btn_credenciales: "Iniciar sesion para SAP"
            }

            for widget, text in tooltips.items():
                self._add_tooltip(widget, text)

    def _add_tooltip(self, widget, text):
            tooltip = tk.Toplevel(widget)
            tooltip.withdraw()
            tooltip.overrideredirect(True)
            label = tk.Label(tooltip, text=text, background="white", relief="solid", borderwidth=1)
            label.pack()
            def enter(event):
                tooltip.geometry(f"+{event.x_root+10}+{event.y_root+10}")
                tooltip.deiconify()
            def leave(event):
                tooltip.withdraw()
            widget.bind("<Enter>", enter)
            widget.bind("<Leave>", leave)

    def abrir_app_mainboard(self):
        if hasattr(self, "_mainboard_proc") and self._mainboard_proc.poll() is None:
            return

        try:
            #! Ejecutar como módulo desde la raíz
            self._mainboard_proc = subprocess.Popen([
                sys.executable,
                "-m", "backend.UI.motherboard_app"
            ])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la app:\n{e}")
            
    def limpiar_datos(self):
        self.log.config(state="normal")
        self.log.delete("1.0", tk.END)
        self.log.config(state="disabled")
        self.log_msg("[INFO] Limpieza iniciada", "INFO")

        if not self.excel_path.get():
            messagebox.showwarning("Atención", "Selecciona un Excel primero")
            return

        if not self.cargar_excel_datos(ignorar_process=True):
            return

        respuesta = messagebox.askyesno(
            "Procesamiento de Excel",
            "¿Deseas procesar los archivos Excel?"
        )
        if not respuesta:
            return

        try:
            from backend.utils.clean_excel_p2 import procesar_archivo_principal_mainboard_2
        except ImportError as e:
            self.log_msg(f"[ERROR] No se pudo importar procesar_mainboard_P2.py: {e}", "ERROR")
            return

        folder = MAINBOARD_2_FILES_FOLDER
        if not os.path.exists(folder):
            self.log_msg(f"[ERROR] La carpeta {folder} no existe")
            return

        carpeta_final = os.path.join(folder, "ARCHIVOS_FINALES")
        os.makedirs(carpeta_final, exist_ok=True)

        archivos = [
            f for f in os.listdir(folder)
            if os.path.isfile(os.path.join(folder, f)) and f.lower().endswith(".xlsx")
        ]
        archivos.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)))

        archivos_procesados = cargar_archivos_procesados()

        for i, f in enumerate(archivos):
            if f in archivos_procesados:
                self.log_msg(f"[INFO] Archivo ya procesado, se omite: {f}\n", "INFO")
                continue

            ruta_excel = os.path.join(folder, f)
            salida_excel = os.path.join(carpeta_final, f"MB-BMM-{f}")

            try:
                self.log_msg(f"[INFO] Procesando archivo: {f}", "OK")
                internal_model = self.internal_models[i] if i < len(self.internal_models) else ""
                plantas = self.plantas[i] if i < len(self.plantas) else ""

                procesar_archivo_principal_mainboard_2(
                ruta_excel_principal=ruta_excel,
                ruta_salida_principal=salida_excel,
                internal_model=internal_model,
                plantas=plantas,
                df_no_procesadas=self.df_no_procesadas
            )

                guardar_archivo_procesado(f)
                self.log_msg(f"[OK] Archivo procesado: PROCESADO_{f}\n")

            except Exception as e:
                self.log_msg(f"[ERROR] No se pudo procesar {f}: {e}", "ERROR")

        self.log_msg("[INFO] Todos los archivos nuevos han sido procesados")
        
        #! Mensaje de limpieza finalizado 
        messagebox.showinfo(
            "Limpieza finalizada",
            "Los archivos han sido limpiados y procesados correctamente 🧹"
        )
        
    #!  VALIDACIÓN DE BOTONES 
    def verificar_habilitar_botones(self):
        cred = cargar_credenciales()

        cred_ok = all(cred.get(k) for k in [
            "SAP_SYSTEM_NAME",
            "SAP_USER",
            "SAP_PASSWORD"
        ])

        excel = self.excel_path.get().strip()

        if not cred_ok:
            #! Bloquear todo si no hay credenciales
            self.btn_mainboard.config(state="disabled")
            self.btn_procesar.config(state="disabled")
            self.btn_limpiar.config(state="disabled")
            self.btn_open.config(state="disabled")

            self.status.set("Estado: Ingrese credenciales SAP 🔐")

        else:
            #! Habilitar botones
            self.btn_mainboard.config(state="normal")
            self.btn_limpiar.config(state="normal")

            if excel:
                self.btn_procesar.config(state="normal")
            else:
                self.btn_procesar.config(state="disabled")

        #! Resultados solo se habilita al final
        self.btn_open.config(state="disabled")


    #!  CREDENCIALES 
    def abrir_credenciales(self):
        cred = cargar_credenciales()
        win = tk.Toplevel(self.root)
        win.title("Credenciales SAP")
        win.geometry("320x240")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        try:
            img = Image.open("backend/IMG/logo.png")
            img = img.resize((256, 256))
            icon = ImageTk.PhotoImage(img)
            win.iconphoto(True, icon)
        except Exception as e:
                print(f"No se pudo cambiar el icono: {e}")

        ttk.Label(win, text="Sistema SAP").pack(pady=(12, 0))
        sistema = tk.StringVar(value=cred.get("SAP_SYSTEM_NAME", ""))
        ttk.Entry(win, textvariable=sistema).pack(fill="x", padx=20)

        ttk.Label(win, text="Usuario").pack(pady=(10, 0))
        usuario = tk.StringVar(value=cred.get("SAP_USER", ""))
        ttk.Entry(win, textvariable=usuario).pack(fill="x", padx=20)

        ttk.Label(win, text="Contraseña").pack(pady=(10, 0))
        password = tk.StringVar(value=cred.get("SAP_PASSWORD", ""))
        ttk.Entry(win, textvariable=password, show="*").pack(fill="x", padx=20)

        def guardar():
            if not sistema.get() or not usuario.get() or not password.get():
                messagebox.showwarning("Atención", "Todos los campos son obligatorios")
                return
            guardar_credenciales({
                "SAP_SYSTEM_NAME": sistema.get().strip(),
                "SAP_USER": usuario.get().strip(),
                "SAP_PASSWORD": password.get()
            })
            messagebox.showinfo("OK", "Credenciales guardadas correctamente")
            win.destroy()
            self.verificar_habilitar_botones()
        ttk.Button(win, text="Guardar", command=guardar).pack(pady=18)

    #!  LOG 
    def log_msg(self, msg, tag="INFO"):
        self.log.config(state="normal")
        self.log.insert(tk.END, msg + "\n", tag)
        self.log.see(tk.END)
        self.log.config(state="disabled")
        self.root.update()

    #!  ESTADO DE LA APP 
    def set_status(self, msg, animar=False):
        self.animando = False
        self.anim_dots = 0
        if animar:
            self.animando = True
            self.animar_estado(msg)
        else:
            self.status.set(f"Estado: {msg}")
        self.root.update()

    def animar_estado(self, texto):
        if not self.animando:
            return
        self.anim_dots = (self.anim_dots + 1) % 4
        self.status.set(f"Estado: {texto}{'.' * self.anim_dots}")
        self.root.after(500, lambda: self.animar_estado(texto))

    def actualizar_tiempo(self):
        if not self.start_time:
            return
        m, s = divmod(int(time.time() - self.start_time), 60)
        base = self.status.get().split(" | ")[0]
        self.status.set(f"{base} | ⏱ {m:02d}:{s:02d}")
        if self.btn_procesar["state"] == "disabled":
            self.root.after(1000, self.actualizar_tiempo)

    #! UI
    def seleccionar_excel(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*")])
        if f:
            self.excel_path.set(f)

    def abrir_resultados(self):
        path = os.path.abspath(MAINBOARD_2_FILES_FOLDER)
        if os.path.exists(path):
            os.startfile(path)
            

    #! FLUJO 
    def iniciar(self):
        #! Validar credenciales antes de iniciar
        cred = cargar_credenciales()
        if not cred.get("SAP_SYSTEM_NAME") or not cred.get("SAP_USER") or not cred.get("SAP_PASSWORD"):
            messagebox.showerror(
                "Credenciales incompletas",
                "No se han iniciado sesionn en SAP.\n"
                "Por favor ve a 🔐 usuario y contraseña e ingrésalas antes de continuar."
            )
            return  #! No continuar hasta que estén completas

        #! Validar que haya un Excel seleccionado
        if not self.excel_path.get():
            messagebox.showwarning("Atención", "Selecciona un Excel")
            return

        #! Deshabilitar botones y preparar progreso
        self.btn_procesar.config(state="disabled")
        self.btn_open.config(state="disabled")
        self.progress["value"] = 0

        self.start_time = time.time()
        self.actualizar_tiempo()

        self.log_msg("[INFO] Automatización iniciada")
        self.set_status("Cargando Excel", animar=True)
        self.root.after(100, self.flujo_procesar)
        
        
    def flujo_procesar(self):
        if not self.cargar_excel_datos():
            self.btn_procesar.config(state="normal")
            return

        self.set_status("Conectando a SAP", animar=True)

        try:
            self.session = abrir_sap_y_login()
        except Exception as e:
            self.log_msg(f"[ERROR] {e}", "ERROR")
            self.btn_procesar.config(state="normal")
            self.animando = False
            return

        self.animando = False
        self.log_msg("[OK] Conectado a SAP", "OK")
        self.idx = 0
        self.root.after(200, self.procesar_modelo)
            
    def cargar_excel_datos(self, ignorar_process=False):
        try:
            df = pd.read_excel(self.excel_path.get())
            df.columns = df.columns.str.strip().str.upper()

            columnas_requeridas = ["MATERIAL", "PLANT", "ALTBOM", "INTERNAL MODEL", "PROCESS", "MAINBOARD PART NUMBER"]
            faltantes = [c for c in columnas_requeridas if c not in df.columns]
            if faltantes:
                raise ValueError(f"No se encontraron las columnas: {faltantes}")

            df["PROCESS"] = df["PROCESS"].apply(lambda x: True if str(x).upper() == "TRUE" else False)

            #! Si es limpieza → NO filtrar
            if ignorar_process:
                df_filtrado = df
            else:
                df_filtrado = df[df["PROCESS"] == False]
                if df_filtrado.empty:
                    raise ValueError("No hay filas nuevas para procesar (todas son TRUE)")

            def limpiar_columna(nombre_columna):
                return (
                    df_filtrado[nombre_columna]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .str.replace(r"\.0$", "", regex=True)
                    .tolist()
                )

            self.modelos = limpiar_columna("MATERIAL")
            self.plantas = limpiar_columna("PLANT")
            self.altboms = limpiar_columna("ALTBOM")
            self.internal_models = limpiar_columna("INTERNAL MODEL")
            self.mainboard_numbers = df_filtrado["MAINBOARD PART NUMBER"].astype(str).str.strip().tolist()

            self.df_no_procesadas = df_filtrado

            self.log_msg("Excel cargado correctamente", "OK")
            return True

        except Exception as e:
            self.log_msg(f"[ERROR] {e}", "ERROR")
            return False

    def procesar_modelo(self):
        total = len(self.modelos)

        if total == 0:
            self.log_msg("[ERROR] No hay materiales para procesar", "ERROR")
            self.btn_procesar.config(state="normal")
            return

        if self.idx >= total:
            self.log_msg("\n[INFO] Iniciando procesamiento de las motherboards y mainboards\n")
            self.guardar_excel_final()
            self.set_status("Finalizado ✅")
            self.progress["value"] = 100
            self.btn_open.config(state="normal")
            self.btn_procesar.config(state="normal")

            #! Mensaje de proceso compeltado
            messagebox.showinfo(
                "Proceso finalizado",
                "El procesamiento de los 1TE ha terminado correctamente ✅"
                "y el Excel se ha actualido correctamente ✅"
            )

            #! Actualizar solo la columna PROCESS en Excel usando openpyxl
            try:
                wb = load_workbook(self.excel_path.get())
                ws = wb.active

                #! Buscar la columna "PROCESS" (mayúsculas por seguridad)
                col_process = None
                col_material = None
                for i, cell in enumerate(ws[1], start=1):
                    if str(cell.value).strip().upper() == "PROCESS":
                        col_process = i
                    if str(cell.value).strip().upper() == "MATERIAL":
                        col_material = i

                if col_process is None or col_material is None:
                    raise ValueError("No se encontraron las columnas 'MATERIAL' o 'PROCESS' en el Excel")

                #! Actualizar filas procesadas
                materiales_procesados = [str(m).strip() for m in self.materiales_procesados_ok]
                for row in ws.iter_rows(min_row=2):
                    material = str(row[col_material - 1].value).strip()
                    if material in materiales_procesados:
                        row[col_process - 1].value = True  #! Actualizamos solo el valor

                wb.save(self.excel_path.get())
                self.log_msg("[OK] Excel actualizado", "OK")

            except Exception as e:
                self.log_msg(f"[ERROR] No se pudo actualizar el Excel: {e}", "ERROR")

            return
        
        internal_models = self.internal_models[self.idx]
        modelo = self.modelos[self.idx]
        self.set_status(f"Modelo {self.idx + 1}/{total}")
        self.log_msg(f"\n▶ Modelo {self.idx + 1}/{total}: {modelo}", "OK")

        try:
            self.log_msg("  • Ejecutando CS11...\n", "INFO")

            resultados = ejecutar_cs11(
                self.session,
                material=modelo,
                uso=FILTRO,
                altboms=self.altboms[self.idx],
                plantas=self.plantas[self.idx]
            )

            if not resultados:
                self.log_msg(f"[INFO] No se encontraron plantas para {modelo}")

            for planta, _ in resultados:
                self.log_msg(f"  • Planta {planta}: exportando BOM")

                ruta_xls = exportar_bom_a_xls(
                    self.session,
                    modelo,
                    mainboard=False
                )

                self.log_msg("    ✓ BOM exportado")

                fecha = datetime.now().strftime("%Y-%m-%d")
                nombre_base = os.path.splitext(os.path.basename(ruta_xls))[0]
                nombre_base = re.sub(r'[\\/*?:"<>|]', "_", nombre_base)
                nombre_base = re.sub(r'^(?:\d+-)+', '', nombre_base)

                altboms = self.altboms[self.idx]

                ruta_xlsx = os.path.join(
                    MODEL_FILES_FOLDER,
                    f"{fecha}-{nombre_base}-ALTBOM{altboms}.xlsx"
                )

                if os.path.exists(ruta_xlsx):
                    os.remove(ruta_xlsx)

                convertir_xls_a_xlsx(ruta_xls, ruta_xlsx)

                self.log_msg("    ✓ Convertido a XLSX")

                self.log_msg("    • Buscando motherboard", "OK")

                df_modelo = extract_descripcion_numbers(
                    ruta_xlsx,
                    internal_models,
                    DESCRIPCIONES
                )

                if df_modelo.empty:
                    self.log_msg(f"[ERROR] No se encontro motherboard para {modelo}")

                else:
                    df_modelo["Modelo"] = modelo
                    df_modelo["Planta"] = planta
                    self.df_todos = pd.concat([self.df_todos, df_modelo], ignore_index=True)

        except Exception as e:
            self.log_msg(f"[ERROR] {e}", "ERROR")

        else:
            #! solo si terminó correctamente
            self.materiales_procesados_ok.append(modelo)
                
        #! Incrementar índice y continuar
        self.idx += 1
        self.root.after(200, self.procesar_modelo)
    def guardar_excel_final(self):
        self.set_status("Procesando mainboards")

        for folder in [
            BASE_BOM_FOLDER, MODEL_FILES_FOLDER, MAINBOARD_1_FILES_FOLDER,
            MAINBOARD_2_FILES_FOLDER,
        ]:
            os.makedirs(folder, exist_ok=True)

        for _, row in self.df_todos.iterrows():
            number = str(row["Number"]).strip().replace(".0", "")
            modelo = str(row["Modelo"]).strip()

            if any(number in f for f in os.listdir(MAINBOARD_1_FILES_FOLDER)):
                continue

            try:
                ruta_xls = None

                for planta in set(self.plantas):
                    ruta_xls = procesar_number(
                        session=self.session,
                        number=number,
                        planta=planta,
                        uso=FILTRO
                    )
                    if ruta_xls:
                        break

                if not ruta_xls:
                    continue

                ruta_xlsx = os.path.join(
                    MAINBOARD_1_FILES_FOLDER,
                    re.sub(r'[\\/*?:"<>|]', "_", os.path.basename(ruta_xls).rsplit(".", 1)[0]) + ".xlsx"
                )

                convertir_xls_a_xlsx(ruta_xls, ruta_xlsx)
                limpiar_excel_mainboard(ruta_xlsx)

                #! MAINBOARD 1
                try:
                    actualizar_excel_mainboard_1(
                        self.excel_path.get(),
                        modelo,
                        [number]
                    )
                except Exception as e:
                    self.log_msg(f"[ERROR] No se pudo actualizar Excel mainboard 1: {e}", "ERROR")

                #! MAINBOARD 2
                materiales_detectados = []

                for planta in set(self.plantas):
                    material = procesar_material_desde_mainboard(
                        session=self.session,
                        ruta_mainboard_xlsx=ruta_xlsx,
                        uso=FILTRO,
                        planta=planta
                    )

                    if material:
                        materiales_detectados.append(str(material).strip())
                        
                materiales_detectados = list(set(materiales_detectados))

                print("Materiales detectados:", materiales_detectados)

                if materiales_detectados:  #! evitar escribir vacío
                    try:
                        actualizar_excel_mainboard_2(
                            self.excel_path.get(),
                            modelo,
                            materiales_detectados
                        )
                    except Exception as e:
                        self.log_msg(f"[ERROR] No se pudo actualizar Excel mainboard 2: {e}", "ERROR")

            except Exception as e:
                self.log_msg(f"[ERROR] Mainboard {number}: {e}", "ERROR")

            #! limpiar .xls basura
            for folder in [
                MAINBOARD_1_FILES_FOLDER,
                MAINBOARD_2_FILES_FOLDER,
                MODEL_FILES_FOLDER
            ]:
                for f in os.listdir(folder):
                    ruta = os.path.join(folder, f)
                    if os.path.isfile(ruta) and f.lower().endswith(".xls"):
                        try:
                            os.remove(ruta)
                        except Exception as e:
                            self.log_msg(f"[ERROR] No se pudo eliminar {f}: {e}", "ERROR")


if __name__ == "__main__":
    root = tk.Tk()
    app = SAPApp(root)
    
    #! Validación inmediata de credenciales
    cred = cargar_credenciales()
    if not cred.get("SAP_SYSTEM_NAME") or not cred.get("SAP_USER") or not cred.get("SAP_PASSWORD"):
        messagebox.showinfo(
            "Atención",
            "No se han iniciado sesionn en SAP.\n"
            "Ve a 🔐 inicar sesion para habilitar cualquier proceso."
        )

    root.mainloop()