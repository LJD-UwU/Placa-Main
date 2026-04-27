import os
import time
import threading
import logging
import concurrent.futures
import pandas as pd
import tkinter as tk
import xlwings as xw
from datetime import datetime
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import ttk, filedialog, scrolledtext, messagebox

try:
    import pythoncom
    _PYTHONCOM_AVAILABLE = True
except ImportError:
    _PYTHONCOM_AVAILABLE = False

from backend.config.sap_config import FILTRO
from backend.config.sap_login import abrir_sap_y_login
from backend.utils.clean_excel import limpiar_excel_mainboard
from backend.utils.utils_2.xlsx_m2 import convertir_xls_a_xlsx
from backend.modules.Modules_2.procesar_mainboard import actualizar_excel_mainboard
from backend.modules.Modules_2.procesar_motherboard import procesar_numbers_desde_listas
from backend.utils.txt_to_xlsx import (MAINBOARD_1_FILES_FOLDER, MAINBOARD_2_FILES_FOLDER)
from backend.modules.Modules_2.procesar_mainboard import procesar_material_desde_mainboard

#! helpers (SAP GUI y xlwings usan COM en Windows) 

def _coinit():
    """
    Inicializa COM en el hilo actual.
    OBLIGATORIO al inicio de cada hilo secundario que use SAP o xlwings.
    Sin esto: error (-2147417842) RPC_E_WRONG_THREAD.
    """
    if _PYTHONCOM_AVAILABLE:
        pythoncom.CoInitialize()


def _couninit():
    """Libera COM en el hilo actual. Llamar siempre en el finally del hilo."""
    if _PYTHONCOM_AVAILABLE:
        pythoncom.CoUninitialize()


#! Helpers de Excel 

def limpiar_valor(valor):
    if valor is None:
        return ""
    valor = str(valor).strip()
    if valor.endswith(".0"):
        valor = valor[:-2]
    return valor


def abrir_excel_seguro(path_excel):
    """
    Abre Excel con fallback:
    1. openpyxl (rápido, sin COM)
    2. xlwings  (soporta cifrado, requiere COM)
    """
    try:
        wb = load_workbook(path_excel)
        return wb, "openpyxl"
    except Exception:
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False
        try:
            wb_xw = app.books.open(path_excel)
            return (app, wb_xw), "xlwings"
        except Exception as e:
            app.quit()
            raise Exception(f"No se pudo abrir el archivo (ni openpyxl ni xlwings): {e}")


def es_amarillo(celda):
    """Detecta celda amarilla en openpyxl."""
    if not celda.fill:
        return False
    fg = celda.fill.fgColor
    try:
        if fg.rgb and str(fg.rgb).upper().endswith("FFFF00"):
            return True
        if fg.indexed is not None and fg.indexed == 6:
            return True
    except Exception:
        pass
    return False


def leer_filas_amarillas(path_excel):
    """
    Lee las filas con celda amarilla en 'MOTHERBOARD PART NUMBER'.
    Soporta openpyxl y xlwings. No requiere COM propio —
    el llamador debe haber hecho _coinit() si usa xlwings.
    """
    obj, engine = abrir_excel_seguro(path_excel)
    mothers, plants, internals = [], [], []

    try:
        if engine == "openpyxl":
            wb = obj
            ws = wb.active

            headers = {}
            for col in range(1, ws.max_column + 1):
                nombre = ws.cell(row=1, column=col).value
                if nombre:
                    headers[str(nombre).strip().upper()] = col

            required = ["MOTHERBOARD PART NUMBER", "PLANT", "INTERNAL MODEL"]
            for r in required:
                if r not in headers:
                    raise Exception(f"No se encontró la columna: {r}")

            col_mother   = headers["MOTHERBOARD PART NUMBER"]
            col_plant    = headers["PLANT"]
            col_internal = headers["INTERNAL MODEL"]

            for row in range(2, ws.max_row + 1):
                celda = ws.cell(row=row, column=col_mother)
                if es_amarillo(celda) and celda.value:
                    mothers.append(limpiar_valor(celda.value))
                    plants.append(limpiar_valor(ws.cell(row=row, column=col_plant).value))
                    internals.append(limpiar_valor(ws.cell(row=row, column=col_internal).value))

        else:  #! xlwings
            app, wb = obj
            sheet = wb.sheets[0]
            data = sheet.used_range.value
            if not data:
                raise Exception("Excel vacío")

            headers = [str(h).strip().upper() for h in data[0]]

            def col_idx(name):
                if name not in headers:
                    raise Exception(f"No se encontró la columna: {name}")
                return headers.index(name)

            col_mother   = col_idx("MOTHERBOARD PART NUMBER")
            col_plant    = col_idx("PLANT")
            col_internal = col_idx("INTERNAL MODEL")

            for i in range(2, len(data) + 1):
                celda = sheet.cells(i, col_mother + 1)
                color = celda.color
                es_amarillo_xw = color and (
                    color[0] > 200 and color[1] > 200 and color[2] < 100
                )
                if es_amarillo_xw:
                    row = data[i - 1]
                    mother   = limpiar_valor(row[col_mother])
                    plant    = limpiar_valor(row[col_plant])
                    internal = limpiar_valor(row[col_internal])
                    if mother:
                        mothers.append(mother)
                        plants.append(plant)
                        internals.append(internal)

    finally:
        if engine == "openpyxl":
            obj.close()
        else:
            app, wb = obj
            try:
                wb.close()
            except Exception:
                pass
            app.quit()

    return mothers, plants, internals


def marcar_procesado(path_excel, mother_name):
    """Colorea la celda de verde cuando la motherboard fue procesada."""
    try:
        app = xw.App(visible=False)
        wb  = app.books.open(path_excel)
        sheet = wb.sheets[0]
        data  = sheet.used_range.value
        headers = [str(h).strip().upper() for h in data[0]]
        col_mother = headers.index("MOTHERBOARD PART NUMBER")

        for i in range(2, len(data) + 1):
            celda = sheet.cells(i, col_mother + 1)
            if limpiar_valor(celda.value) == limpiar_valor(mother_name):
                celda.color = (198, 224, 180)
                break

        wb.save()
        wb.close()
        app.quit()

    except Exception as e:
        logging.error(f"marcar_procesado({mother_name}): {e}")


def eliminar_xls_carpeta(carpeta):
    """Elimina todos los .xls temporales de una carpeta."""
    for f in os.listdir(carpeta):
        if f.lower().endswith(".xls"):
            ruta = os.path.join(carpeta, f)
            try:
                os.remove(ruta)
                logging.info(f"Eliminado: {ruta}")
            except Exception as e:
                logging.error(f"No se pudo eliminar {ruta}: {e}")


#! Aplicación 

class MainboardApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("MBAutomator - Motherboard")
        self.root.geometry("410x420")
        self.root.resizable(False, False)

        try:
            img  = Image.open("backend/IMG/logo.png").resize((256, 256))
            icon = ImageTk.PhotoImage(img)
            self.root.iconphoto(True, icon)
        except Exception:
            pass

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Title.TLabel", font=("Segoe UI", 14, "bold"))

        ttk.Label(root, text="Automatización SAP", style="Title.TLabel").pack(pady=(8, 0))
        ttk.Label(root, text="Procesamiento de Excel para Motherboard", foreground="gray").pack(pady=(0, 6))

        main = ttk.Frame(root, padding=6)
        main.pack(fill="both", expand=True)

        fila_file = ttk.Frame(main)
        fila_file.pack(fill="x", pady=4)
        self.excel_path = tk.StringVar()
        ttk.Entry(fila_file, textvariable=self.excel_path).pack(side="left", fill="x", expand=True)
        ttk.Button(fila_file, text="📂", width=3, command=self.seleccionar_excel).pack(side="left", padx=4)

        fila_btn = ttk.Frame(main)
        fila_btn.pack(pady=4)
        self.btn_procesar = ttk.Button(fila_btn, text="▶ Procesar", command=self.iniciar_procesamiento)
        self.btn_procesar.pack(side="left", padx=4)

        frame_log = ttk.LabelFrame(main, text="CONSOLA")
        frame_log.pack(fill="both", expand=True, pady=(6, 0))
        self.log = scrolledtext.ScrolledText(frame_log, height=10, font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=5, pady=5)
        self.log.config(state="disabled")
        self.log.tag_config("INFO",  foreground="blue")
        self.log.tag_config("OK",    foreground="green")
        self.log.tag_config("ERROR", foreground="red")

        self.status = tk.StringVar(value="Estado: Listo")
        ttk.Label(root, textvariable=self.status, anchor="w").pack(fill="x", side="bottom", padx=6, pady=4)

        self.session       = None
        self.excel_paths   = []
        self.mother        = []
        self.plants        = []
        self.internal_models = []

    #! LOGGING THREAD-SAFE 

    def log_msg(self, msg: str, tag: str = "INFO"):
        """
        Siempre thread-safe: encola en el hilo principal via root.after().
        NUNCA llama root.update() desde un hilo secundario.
        """
        logging.debug(f"[{tag}] {msg}")
        self.root.after(0, self._log_msg_ui, msg, tag)

    def _log_msg_ui(self, msg: str, tag: str):
        """Solo llamado desde el hilo principal."""
        self.log.config(state="normal")
        self.log.insert(tk.END, msg + "\n", tag)
        self.log.see(tk.END)
        self.log.config(state="disabled")

    def set_status(self, msg: str):
        """Thread-safe: actualiza la etiqueta de estado."""
        self.root.after(0, lambda: self.status.set(f"Estado: {msg}"))

    #! UI 

    def seleccionar_excel(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel", "*.xlsx")])
        if files:
            self.excel_paths = list(files)
            self.log_msg(f"[OK] {len(files)} archivo(s) seleccionado(s)", "OK")

    #! INICIO 

    def iniciar_procesamiento(self):
        if not self.excel_paths:
            messagebox.showwarning("Atención", "Selecciona al menos un archivo Excel primero.")
            return

        #! Deshabilitar botón para evitar doble clic
        self.btn_procesar.config(state="disabled")
        self.log_msg("[INFO] Automatización iniciada\n", "INFO")
        self.set_status("Procesando...")

        #! Todo el trabajo pesado en un hilo secundario
        threading.Thread(target=self._worker, daemon=True).start()

    #! WORKER 

    def _worker(self):
        """
        Corre íntegramente en un hilo secundario.
        CoInitialize/CoUninitialize envuelven TODA la interacción COM
        (SAP GUI + xlwings).
        """
        _coinit()
        try:
            self._conectar_sap()
            if not self.session:
                return

            os.makedirs(MAINBOARD_1_FILES_FOLDER, exist_ok=True)

            for excel_file in self.excel_paths:
                self._procesar_archivo(excel_file)

            #! Finalización
            self.log_msg("[INFO] Todas las motherboards se procesaron", "INFO")
            self.log_msg("[OK] Proceso completo", "OK")

            #! messagebox siempre en el hilo principal
            self.root.after(0, self._on_proceso_completado)

        except Exception as e:
            self.log_msg(f"[ERROR] Error fatal: {e}", "ERROR")
            logging.exception("Error fatal en _worker")
        finally:
            _couninit()
            #! Re-habilitar botón siempre, incluso si hubo error
            self.root.after(0, lambda: self.btn_procesar.config(state="normal"))
            self.set_status("Listo")

    def _conectar_sap(self):
        """Conecta a SAP. Corre en hilo secundario (COM ya inicializado)."""
        if not self.session:
            try:
                self.log_msg("[INFO] Conectando a SAP...", "INFO")
                self.session = abrir_sap_y_login()
                self.log_msg("[OK] Conectado a SAP", "OK")
            except Exception as e:
                self.log_msg(f"[ERROR] Login SAP: {e}", "ERROR")
                logging.exception("Error login SAP")
                self.session = None

    def _procesar_archivo(self, excel_file: str):
        """Procesa un archivo Excel completo. Corre en hilo secundario."""
        eliminar_xls_carpeta(MAINBOARD_1_FILES_FOLDER)
        eliminar_xls_carpeta(MAINBOARD_2_FILES_FOLDER)

        archivo_nombre = os.path.basename(excel_file)
        self.log_msg(f"▶ Archivo: {archivo_nombre}", "INFO")
        self.log_msg("  • Leyendo Excel", "INFO")

        try:
            mothers, plants, internals = leer_filas_amarillas(excel_file)
        except Exception as e:
            self.log_msg(f"    [ERROR] No se pudieron leer filas amarillas: {e}", "ERROR")
            return

        total = len(mothers)
        if total == 0:
            self.log_msg("    ⚠ No hay filas amarillas → archivo omitido", "ERROR")
            return

        self.log_msg(f"    ✓ {total} Motherboard(s) a procesar\n", "OK")

        for idx, (mother, plant) in enumerate(zip(mothers, plants), start=1):
            self.log_msg(f"▶ Motherboard {idx}/{total}: {mother}", "INFO")
            self.log_msg(f"  • Planta: {plant}", "INFO")
            self.set_status(f"Motherboard {idx}/{total}: {mother}")

            fecha        = datetime.now().strftime("%Y-%m-%d")
            excel_salida = os.path.join(MAINBOARD_1_FILES_FOLDER, f"{fecha}-{mother}.xlsx")

            try:
                procesar_numbers_desde_listas(
                    session=self.session,
                    mother_list=[mother],
                    plant_list=[plant],
                    excel_output=excel_salida,
                    capid=FILTRO,
                )

                ruta_xls = os.path.join(MAINBOARD_1_FILES_FOLDER, f"{fecha}-{mother}.XLS")

                #! Esperar hasta 20 s a que SAP genere el XLS
                for _ in range(20):
                    if os.path.exists(ruta_xls):
                        break
                    time.sleep(1)
                else:
                    raise Exception("El XLS no fue generado por SAP en 20 s")

                convertir_xls_a_xlsx(ruta_xls, excel_salida)
                limpiar_excel_mainboard(excel_salida)

                resultado = procesar_material_desde_mainboard(
                    session=self.session,
                    ruta_mainboard_xlsx=excel_salida,
                    uso=FILTRO,
                    planta=plant,
                    mother=mother,
                )

                if resultado is None:
                    self.log_msg(f"    [ERROR] No se pudo generar BOM para {mother}\n", "ERROR")
                    continue

                ruta_xlsx, material_detectado, descripcion_detectada = resultado

                try:
                    actualizar_excel_mainboard(
                        mother,
                        [material_detectado],
                        ruta_excel=excel_file,
                        descripcion=descripcion_detectada,
                    )
                    self.log_msg(f"    ✓ Excel actualizado con materiales de {mother}", "OK")
                except Exception as e:
                    self.log_msg(f"    [ERROR] No se pudo actualizar Excel: {e}", "ERROR")

                #! marcar_procesado también usa xlwings → COM ya inicializado en este hilo
                marcar_procesado(excel_file, mother)
                self.log_msg(f"    ✓ BOM generado correctamente\n", "OK")

            except Exception as e:
                self.log_msg(f"    [ERROR] Error procesando {mother}: {e}\n", "ERROR")
                logging.exception(f"Error procesando motherboard {mother}")

    def _on_proceso_completado(self):
        """
        Llamado desde root.after — corre en el hilo principal.
        Muestra mensajes y hace limpieza final.
        """
        messagebox.showinfo(
            "Proceso terminado",
            "El proceso de Motherboard ha finalizado correctamente",
        )

        #! Limpieza final de XLS (no hace COM — solo os.remove)
        self.log_msg("[INFO] Limpiando archivos XLS antes de cerrar...", "INFO")
        try:
            eliminar_xls_carpeta(MAINBOARD_1_FILES_FOLDER)
            eliminar_xls_carpeta(MAINBOARD_2_FILES_FOLDER)
            self.log_msg("[OK] Archivos XLS eliminados", "OK")
        except Exception as e:
            self.log_msg(f"[ERROR] No se pudieron eliminar los XLS: {e}", "ERROR")

        messagebox.showinfo(
            "Siguiente paso",
            "Cargar archivo procesado en la ventana principal y dar clic en el botón 'Limpiar'",
        )

        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app  = MainboardApp(root)
    root.mainloop()