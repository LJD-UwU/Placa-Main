import os
import time
import openpyxl
from backend.utils.txt_to_xlsx import MAINBOARD_1_FILES_FOLDER


def mover_columnas_por_nombre(ws, columnas_a_mover, antes_de):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_index = {str(name).strip(): i + 1 for i, name in enumerate(headers) if name}

    for col in columnas_a_mover + [antes_de]:
        if col not in col_index:
            print(f"[WARNING] Columna no encontrada: {col}")
            return

    destino = col_index[antes_de]

    data_cols = []
    for col in columnas_a_mover:
        idx = col_index[col]
        data = [ws.cell(row=r, column=idx).value for r in range(1, ws.max_row + 1)]
        data_cols.append((col, data))

    for col in sorted(columnas_a_mover, key=lambda x: col_index[x], reverse=True):
        ws.delete_cols(col_index[col])

    for i, (nombre, data) in enumerate(data_cols):
        ws.insert_cols(destino + i)
        for r, val in enumerate(data, start=1):
            ws.cell(row=r, column=destino + i).value = val


def limpiar_excel_mainboard(ruta_xlsx: str):
    wb = openpyxl.load_workbook(ruta_xlsx)
    ws = wb.active

    #! LIMPIEZA BASE
    ws.delete_cols(1,2)
    ws.delete_cols(7)
    ws.delete_cols(10)
    ws.delete_rows(1, 9)

    #! MOVER COLUMNAS (tu lógica integrada)
    mover_columnas_por_nombre(
        ws,
        columnas_a_mover=["组件数量", "Un"],
        antes_de="项目文本行 1"
    )

    #! HEADERS
    headers = [
        "LEVEL", "ITEM", "MATERIAL",
        "DESCRIPTION IN CHINESE", "DESCRIPTION IN ENGLISH",
        "QTY", "UN", "LINE 1", "LINE 2", "SORT STRING"
    ]

    ws.insert_cols(1)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    wb.save(ruta_xlsx)


def limpiar_todos_los_mainboard():
    for archivo in os.listdir(MAINBOARD_1_FILES_FOLDER):
        if archivo.lower().endswith(".xlsx"):
            ruta = os.path.join(MAINBOARD_1_FILES_FOLDER, archivo)

            try:
                limpiar_excel_mainboard(ruta)
                print(f"[OK] Limpio: {archivo}\n")
                time.sleep(1)
            except Exception as e:
                print(f"[ERROR] {archivo} → {e}")