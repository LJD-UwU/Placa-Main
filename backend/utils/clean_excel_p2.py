import os
import re
import time
import openpyxl
import pandas as pd
import xlwings as xw
from glob import glob
from openpyxl.styles import PatternFill
from backend.config.sap_config import EXTRAER_ARCHIVO
from openpyxl.styles import PatternFill, Font, Alignment

def limpiar_valor(valor):
    if pd.isna(valor):
        return ""
    valor = str(valor).strip()
    if valor.endswith(".0"):
        valor = valor[:-2]
    return valor


def contiene_chino(texto):
    if not texto:
        return False

    texto = str(texto)
    texto = texto.replace("\xa0", "").replace("\u200b", "").strip()

    return any('\u3400' <= c <= '\u9FFF' for c in texto)


def extraer_codigo_pcb(texto, siguiente_celda=None):
    if isinstance(texto, str) and re.search(r'[A-Za-z].*\d|\d.*[A-Za-z]', texto):
        if siguiente_celda:
            match = re.search(r"\.(\d+)(\\|$)", str(siguiente_celda))
            if match:
                return match.group(1)
    return None


def colorear_chino(ws):
    """
    ✅ COMPORTAMIENTO FINAL:
    - Detecta si LINE 1 o LINE 2 contienen caracteres chinos
    - Si contiene chino: 
      * Colorea TODA LA FILA en amarillo
      * ELIMINA SOLO el contenido con chino (vacía la celda)
      * Conserva todo lo demás de la fila
    
    Ejemplos:
    - "U11涂完硅脂后安装" en LINE 1 → Fila amarilla, celda vacía, resto conservado
    - "点胶固定,上件XP1005" en LINE 2 → Fila amarilla, celda vacía, resto conservado
    - "SMT-123" (sin chino) → Sin cambios
    """
    amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    col_indices = {
        str(ws.cell(row=1, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
    }

    col_line1 = col_indices.get("LINE 1")
    col_line2 = col_indices.get("LINE 2")

    for row in range(2, ws.max_row + 1):
        tiene_chino = False
        celdas_con_chino = []
        
        #! Verificar si LINE 1 o LINE 2 contienen chino
        for col in [col_line1, col_line2]:
            if col:
                valor = ws.cell(row=row, column=col).value
                if valor and contiene_chino(valor):
                    tiene_chino = True
                    celdas_con_chino.append(col)
        
        #! Si la fila tiene chino
        if tiene_chino:
            #! Colorear TODA la fila en amarillo
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = amarillo
            
            #! SOLO vaciar las celdas que contienen chino
            for col in celdas_con_chino:
                ws.cell(row=row, column=col).value = None


def aplicar_logica_x(ws):
    col_indices = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_item = col_indices.get("ITEM")
    col_level = col_indices.get("LEVEL")
    if not col_item or not col_level:
        return

    filas_protegidas = {2, 3, 4}
    bold_font = Font(bold=True)

    for row in range(2, ws.max_row + 1):
        if row in filas_protegidas:
            continue
        val = ws.cell(row=row, column=col_item).value
        if val is None or str(val).strip() == "":
            ws.cell(row=row, column=col_item).value = "X"
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).font = bold_font

    contador = 10
    for row in range(2, ws.max_row + 1):
        if row in filas_protegidas:
            continue
        val = ws.cell(row=row, column=col_item).value
        if val == "X":
            contador = 10
            continue
        ws.cell(row=row, column=col_item).value = str(contador)
        contador += 10

    nivel_actual = 1
    for row in range(2, ws.max_row + 1):
        if row in filas_protegidas:
            continue
        val = ws.cell(row=row, column=col_item).value
        if val == "X":
            nivel_actual += 1
        else:
            ws.cell(row=row, column=col_level).value = nivel_actual + 1

    for row in range(3, ws.max_row + 1):
        if ws.cell(row=row, column=col_level).value in (None, ""):
            ws.cell(row=row, column=col_level).value = ws.cell(row=row - 1, column=col_level).value
            
#!  SUBMATERIALES 
def agregar_submateriales(df_main, ws):
    """
    Agrega submateriales de BOM y manuales dentro del LEVEL 2 antes de la X correspondiente.
    Aplica color gris y fuente Calibri 11 sin negrita a todos los submateriales.
    """
    gris_submaterial = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fuente_normal = Font(name="Calibri", size=11, bold=False)

    #! Extraer códigos PCB
    df_main["PCB_CODE"] = df_main.apply(
        lambda row: extraer_codigo_pcb(row["MATERIAL"], row["DESCRIPTION IN CHINESE"]), axis=1
    )
    lista_pcb = df_main["PCB_CODE"].dropna().tolist()
    if not lista_pcb:
        df_main.drop(columns=["PCB_CODE"], inplace=True, errors="ignore")
        return df_main

    #! Cargar archivo más reciente de los submateriales
    archivos = [f for f in glob(os.path.join(EXTRAER_ARCHIVO, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if len(archivos) < 2:
        print("[WARN] No hay suficientes archivos BOM para tomar el anterior al más reciente.")
        df_main.drop(columns=["PCB_CODE"], inplace=True, errors="ignore")
        return df_main

    archivo_bom = sorted(archivos, key=os.path.getmtime)[-1]
    print(f"[INFO] Archivo BOM tomado: {archivo_bom}")

    #! 🔥 LECTURA ROBUSTA (pandas + xlwings fallback)
    try:
        try:
            df_bom = pd.read_excel(archivo_bom)
        except Exception:
            print("[INFO] Pandas no pudo leer el archivo. Intentando con xlwings...")

            app = None
            try:
                app = xw.App(visible=False)
                app.display_alerts = False
                app.screen_updating = False

                wb = app.books.open(archivo_bom)
                sheet = wb.sheets[0]

                data = sheet.used_range.value

                # ! convertir manualmente (más confiable)
                df_bom = pd.DataFrame(data[1:], columns=data[0])

                wb.close()

                df_bom.dropna(how="all", inplace=True)
                df_bom.columns = df_bom.columns.astype(str).str.strip()

            finally:
                if app:
                    app.quit()

    except Exception as e:
        print(f"[ERROR] No se pudo leer {archivo_bom}: {e}")
        df_main.drop(columns=["PCB_CODE"], inplace=True, errors="ignore")
        return df_main
    
    #! Filtrar USE
    df_bom["PCB_clean"] = df_bom["PCB"].apply(limpiar_valor)

    if "USE/NO USE" in df_bom.columns:
        df_bom["USE/NO USE"] = df_bom["USE/NO USE"].astype(str).str.strip().str.upper()
        df_bom = df_bom[df_bom["USE/NO USE"] != "NO USE"]

    #! Filtrar submateriales relacionados con PCB
    cols_interes = ["PCB","Part #","ZH Description","EN Description","QTY","UNIT"]

    lista_pcb = [limpiar_valor(x) for x in lista_pcb]

    df_filtrado = df_bom[df_bom["PCB_clean"].isin(lista_pcb)][cols_interes].reset_index(drop=True)

    df_filtrado["Part #"] = (
    df_filtrado["Part #"]
    .astype(str)
    .str.strip()
    .str.replace(".0", "", regex=False)
    .str.upper()
)

    finales = {"L600022", "1063182"}

    df_finales = df_filtrado[df_filtrado["Part #"].isin(finales)]
    df_normales = df_filtrado[~df_filtrado["Part #"].isin(finales)]

    #! Mapear columnas BOM → Excel
    col_map = {
        "PCB": "ITEM",
        "Part #": "MATERIAL",
        "ZH Description": "DESCRIPTION IN CHINESE",
        "EN Description": "DESCRIPTION IN ENGLISH",
        "QTY": "QTY",
        "UNIT": "UN"
    }

    def mapear_filas(df_sub):
        df_nuevo = pd.DataFrame(columns=df_main.columns)
        for _, fila in df_sub.iterrows():
            nueva = {col_map[col]: fila[col] for col in df_sub.columns if col in col_map}
            df_nuevo = pd.concat([df_nuevo, pd.DataFrame([nueva], columns=df_main.columns)], ignore_index=True)
        df_nuevo["LEVEL"] = 2
        df_nuevo["_SUBMATERIAL"] = True
        return df_nuevo

    df_sub_normales = mapear_filas(df_normales)
    df_sub_finales = mapear_filas(df_finales)

    #!  Filas manuales 
    filas_manuales = [
        {
            "ITEM": "73467",
            "MATERIAL": "L100022",
            "DESCRIPTION IN CHINESE": "",
            "DESCRIPTION IN ENGLISH": "MB BARCODE LABEL (28mm*8mm)",
            "QTY": "1,000",
            "UN": "PC",
            "LEVEL": 2,
            "_SUBMATERIAL": True
        },
        {
            "ITEM": "7353742",
            "MATERIAL": "L600006",
            "DESCRIPTION IN CHINESE": "",
            "DESCRIPTION IN ENGLISH": "RIBBON\\110mm*450m\\LOCAL 556",
            "QTY": "556",
            "UN": "CM",
            "LEVEL": 2,
            "_SUBMATERIAL": True
        }
    ]
    df_manuales = pd.DataFrame(filas_manuales, columns=df_sub_normales.columns)
    df_sub_normales = pd.concat([df_sub_normales, df_manuales], ignore_index=True)

    #!  Insertar submateriales antes de X del bloque LEVEL 2 
    indices_level2 = [i for i, v in enumerate(df_main["LEVEL"]) if v == 2]
    for idx in reversed(indices_level2):
        x_index = None
        for j in range(idx, len(df_main)):
            if df_main.at[j, "ITEM"] == "X" and df_main.at[j, "LEVEL"] == 2:
                x_index = j
                break  
        if x_index is not None:
            df_main = pd.concat([
                df_main.iloc[:x_index],
                df_sub_normales,
                df_main.iloc[x_index:]
            ], ignore_index=True)

            #! Aplicar color y fuente directamente en ws
            for i in range(len(df_sub_normales)):
                fila_excel = x_index + 2 + i
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=fila_excel, column=c).fill = gris_submaterial
                    ws.cell(row=fila_excel, column=c).font = fuente_normal
            break

    #!  Agregar submateriales finales al final 
    df_main = pd.concat([df_main, df_sub_finales], ignore_index=True)
    fila_inicio = len(df_main) - len(df_sub_finales) + 2
    for i in range(len(df_sub_finales)):
        fila_excel = fila_inicio + i
        for c in range(1, ws.max_column + 1):
            ws.cell(row=fila_excel, column=c).fill = gris_submaterial
            ws.cell(row=fila_excel, column=c).font = fuente_normal

    #!  Limpiar columna temporal 
    df_main.drop(columns=["PCB_CODE","_SUBMATERIAL"], inplace=True, errors="ignore")

    return df_main


#!  PROCESO PRINCIPAL 

def mover_columnas_por_nombre(ws, columnas_a_mover, antes_de):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_index = {str(name).strip(): i+1 for i, name in enumerate(headers) if name}

    for col in columnas_a_mover + [antes_de]:
        if col not in col_index:
            print(f"[WARNING] Columna no encontrada: {col}")
            return

    destino = col_index[antes_de]

    data_cols = []
    for col in columnas_a_mover:
        idx = col_index[col]
        data = [ws.cell(row=r, column=idx).value for r in range(1, ws.max_row + 1)]
        data_cols.append((col, data))

    for col in sorted(columnas_a_mover, key=lambda x: col_index[x], reverse=True):
        ws.delete_cols(col_index[col])

    for i, (nombre, data) in enumerate(data_cols):
        ws.insert_cols(destino + i)
        for r, val in enumerate(data, start=1):
            ws.cell(row=r, column=destino + i).value = val


def procesar_archivo_principal_mainboard_2(
    ruta_excel_principal: str,
    ruta_salida_principal: str,
    internal_model: str = "",
    plantas: str = "",
    df_no_procesadas: pd.DataFrame = None
):

    wb = openpyxl.load_workbook(ruta_excel_principal)
    ws = wb.active

    #! LIMPIEZA BASE
    ws.delete_cols(1,2)
    ws.delete_cols(7)
    ws.delete_cols(10)
    ws.delete_rows(1, 9)

    #! LÓGICA: MOVER COLUMNAS
    mover_columnas_por_nombre(
        ws,
        columnas_a_mover=["组件数量", "Un"],
        antes_de="项目文本行 1"
    )

    #! HEADERS 
    headers = [
        "LEVEL", "ITEM", "MATERIAL",
        "DESCRIPTION IN CHINESE", "DESCRIPTION IN ENGLISH",
        "QTY", "UN", "LINE 1", "LINE 2", "SORTSTRNG"
    ]

    ws.insert_cols(1)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    nombre_archivo = os.path.splitext(os.path.basename(ruta_excel_principal))[0]

    ws.insert_rows(2, 2)
    ws["B2"] = "X"
    ws["C2"] = "3TE"
    ws["A3"] = "1"
    ws["B3"] = " "
    ws["C3"] = nombre_archivo
    ws["A4"] = "1"
    ws["B4"] = "X"
    ws["C4"] = nombre_archivo

    aplicar_logica_x(ws)
    time.sleep(0.5)

    df_main = pd.DataFrame(ws.values)
    df_main.columns = df_main.iloc[0]
    df_main = df_main[1:].reset_index(drop=True)

    df_main = agregar_submateriales(df_main, ws)
    time.sleep(0.5)

    filas_protegidas = {0, 1, 2}
    df_main["ITEM"] = df_main["ITEM"].apply(lambda v: str(v).strip() if v else "")
    df_main.loc[~df_main.index.isin(filas_protegidas), "LEVEL"] = 0

    contador_bloque = 10
    for i, val in df_main["ITEM"].items():
        if i in filas_protegidas:
            continue
        if val == "X":
            contador_bloque = 10
            continue
        df_main.at[i, "ITEM"] = str(contador_bloque)
        contador_bloque += 10

    nivel_actual = 1
    for i in range(len(df_main)):
        if i in filas_protegidas:
            continue
        if df_main.at[i, "ITEM"] == "X":
            nivel_actual += 1
        else:
            df_main.at[i, "LEVEL"] = nivel_actual + 1

    for i in range(1, len(df_main)):
        if i in filas_protegidas:
            continue
        if df_main.at[i, "LEVEL"] == 0:
            df_main.at[i, "LEVEL"] = df_main.at[i - 1, "LEVEL"]

    gris_submaterial = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for r_idx, fila in enumerate(df_main.itertuples(index=False), start=2):
        is_submaterial = getattr(fila, "_SUBMATERIAL", False)
        for c_idx, val in enumerate(fila, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val
            if is_submaterial:
                ws.cell(row=r_idx, column=c_idx).fill = gris_submaterial

    #! LIMPIAR PROCESO AI  →  comportamiento según cantidad de "X"─

    #! Contar cuántas "X" hay en la columna ITEM (excluyendo encabezado)
    headers_ws = [cell.value for cell in ws[1]]
    col_item_idx = headers_ws.index("ITEM") + 1 if "ITEM" in headers_ws else None

    cantidad_x = 0
    if col_item_idx:
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_item_idx).value
            if str(val).strip() == "X":
                cantidad_x += 1

    print(f"\n📊 Cantidad de 'X' detectadas : {cantidad_x}")

    #! ── CASO 1 → mantener tal como salió del procesado anterior, sin limpiar AI ──
    if cantidad_x == 3:
            print("✅ 3 X detectadas → Aplicando proceso estandar.\n")
            
    #! ── CASO 2 → Inicia la logica para AI ──
    elif cantidad_x == 4:
        print("🔄 4 X detectadas → Aplicando proceso de AI \n")
        headers = [cell.value for cell in ws[1]]

        col_level = headers.index("LEVEL") + 1
        col_material = headers.index("MATERIAL") + 1

        col_sort = headers.index("SORTSTRNG") + 1 if "SORTSTRNG" in headers else None

        if col_sort is None:
            col_sort = len(headers) + 1
            ws.cell(row=1, column=col_sort, value="SORTSTRNG")

        max_row = ws.max_row
        filas_a_eliminar = set()

        i = 2

        while i <= max_row:
            level_actual = ws.cell(row=i, column=col_level).value
            inicio = i

            while i <= max_row and ws.cell(row=i, column=col_level).value == level_actual:
                i += 1

            fin = i
            tamaño = fin - inicio

            if 3 <= tamaño <= 6:

                #! Buscar padre
                fila_padre = None
                for j in range(inicio - 1, 1, -1):
                    if ws.cell(row=j, column=col_level).value < level_actual:
                        fila_padre = j
                        break

                if fila_padre is None:
                    continue

                #! Obtener MATERIAL del padre
                material_padre = ws.cell(row=fila_padre, column=col_material).value

                #! Reemplazar AI por MATERIAL
                
                for fila in range(inicio, fin - 1):
                    ws.cell(row=fila, column=col_sort).value = material_padre

                #! Última fila del bloque queda vacía
                ws.cell(row=fin - 1, column=col_sort).value = None

                # ! Copiar C y D 
                fila_ultima_bloque = fin - 1
                val_c = ws.cell(row=fila_ultima_bloque, column=3).value
                val_d = ws.cell(row=fila_ultima_bloque, column=4).value

                ws.cell(row=5, column=3).value = val_c
                ws.cell(row=5, column=4).value = val_d

                #! Marcar eliminaciones
                filas_a_eliminar.add(fila_padre)

                #! Eliminar fila debajo del padre
                fila_debajo_padre = fila_padre + 1
                if fila_debajo_padre <= max_row:
                    filas_a_eliminar.add(fila_debajo_padre)

                #! mMntener lógica anterior
                fila_encima_inicio = inicio - 1
                if fila_encima_inicio > 1:
                    filas_a_eliminar.add(fila_encima_inicio)

        #! Eliminar filas 
        for fila in sorted(filas_a_eliminar, reverse=True):
            ws.delete_rows(fila)

    #! ── CASO 3 → proceso SMT A y B ──
    elif cantidad_x == 5:
        print("🔄  5 X detectadas → Aplicando proceso SMT A/B\n")

        fill_color = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") 

        headers = [cell.value for cell in ws[1]]

        col_item = headers.index("ITEM") + 1
        col_material = headers.index("MATERIAL") + 1

        #! 1. Encontrar la 4ta "X"
        contador_x = 0
        fila_cuarta_x = None

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_item).value
            if str(val).strip() == "X":
                contador_x += 1
                if contador_x == 4:
                    fila_cuarta_x = row
                    break

        if not fila_cuarta_x:
            print("⚠️ No se encontró la 4ta X")
            return

        print(f"📍 4ta X encontrada en fila: {fila_cuarta_x}")

        fila_nueva = fila_cuarta_x

        #! 2. Buscar filas con MATERIAL específico
        materiales_objetivo = {"L600022", "1063182"}
        filas_a_mover = []

        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_material).value
            if val and str(val).strip() in materiales_objetivo:
                filas_a_mover.append(row)

        if not filas_a_mover:
            print("⚠️ No se encontraron materiales a mover")
            return

        print(f"📦 Filas a mover: {filas_a_mover}")

        #! 3. Extraer datos
        datos_filas = []
        for fila in filas_a_mover:
            datos = [ws.cell(row=fila, column=col).value for col in range(1, ws.max_column + 1)]
            datos_filas.append(datos)

        #! 4. Eliminar filas originales
        for fila in sorted(filas_a_mover, reverse=True):
            ws.delete_rows(fila)

        #! Ajustar índice
        filas_eliminadas_arriba = sum(1 for f in filas_a_mover if f < fila_nueva)
        fila_nueva -= filas_eliminadas_arriba

        #! 5. Insertar filas movidas (SIN fila extra)
        for i, datos in enumerate(datos_filas):
            ws.insert_rows(fila_nueva + i)

            for col, valor in enumerate(datos, start=1):
                cell = ws.cell(row=fila_nueva + i, column=col)
                cell.value = valor

                if col <= 10:
                    cell.fill = fill_color

        print("✅ Filas movidas correctamente debajo de la 4ta X\n")

    #! ── CASO: cualquier otro valor → advertencia y sin acción ──
    else:
        print(f"⚠️  Cantidad de X no reconocida ({cantidad_x}). No se aplicó ninguna de las logicas.\n")

    #! Limpieza DataFrame
    if "_SUBMATERIAL" in df_main.columns:
        df_main.drop(columns=["_SUBMATERIAL"], inplace=True)

    #! Aplicar negritas
    bold_font = Font(bold=True)
    col_indices = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_item = col_indices.get("ITEM")

    if col_item:
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=col_item).value).strip() == "X":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).font = bold_font

    print("\n✅ Proceso completado correctamente")

    #! RECONSTRUIR ITEM Y LEVEL

    headers = [cell.value for cell in ws[1]]
    col_item = headers.index("ITEM") + 1
    col_level = headers.index("LEVEL") + 1

    filas_protegidas = {2, 3, 4}

    #! RECONSTRUIR ITEM (10,20,30...)
    contador = 10

    for row in range(2, ws.max_row + 1):
        if row in filas_protegidas:
            continue

        val = ws.cell(row=row, column=col_item).value

        if str(val).strip() == "X":
            contador = 10
            continue

        ws.cell(row=row, column=col_item).value = str(contador)
        contador += 10


    #! RECONSTRUIR LEVEL
    nivel_actual = 1

    for row in range(2, ws.max_row + 1):
        if row in filas_protegidas:
            continue

        val = ws.cell(row=row, column=col_item).value

        if str(val).strip() == "X":
            nivel_actual += 1
        else:
            ws.cell(row=row, column=col_level).value = nivel_actual + 1


    #! RELLENAR LEVEL VACÍO
    for row in range(3, ws.max_row + 1):
        if row in filas_protegidas:
            continue

        val = ws.cell(row=row, column=col_level).value

        if val in (None, ""):
            ws.cell(row=row, column=col_level).value = ws.cell(row=row - 1, column=col_level).value

    ws.title = "BOMList"
    ws["A2"] = "0"
    ws["F3"] = "1000"
    ws["J3"] = "HIMEX"
    ws["G3"] = "PC"

    mainboard_num = str(ws["C3"].value).strip().upper()

    df_no_procesadas["MAINBOARD PART NUMBER"] = (
        df_no_procesadas["MAINBOARD PART NUMBER"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    fila_match = df_no_procesadas[
        df_no_procesadas["MAINBOARD PART NUMBER"].str.startswith(mainboard_num)
    ]

    texto_modelo = ""
    if not fila_match.empty:
        texto_modelo = str(fila_match.iloc[0]["INTERNAL MODEL"]).strip()
    else:
        print(f"[WARNING] No se encontró INTERNAL MODEL para {mainboard_num}")

    ws["E3"] = f"MAIN BOARD\\{texto_modelo}\\ROH"
    ws["E4"] = f"MAIN BOARD\\{texto_modelo}\\ROH"
    ws["D3"] = plantas.strip() if plantas else ""

    valor = ws["D5"].value
    if valor and "\\" in valor:
        parte = valor.split("\\", 1)[1]
        ws["E5"] = "MAIN BOARD SMT PART\\" + parte
    else:
        ws["E5"] = "MAIN BOARD SMT PART\\"

    if "BOMHeader" not in wb.sheetnames:
        ws_header = wb.create_sheet("BOMHeader")
        encabezados_header = ["BOMID","MATNR","WERKS","STLAN","STLAL","ZTEXT","BMENG","STKTX"]
        for col, header in enumerate(encabezados_header, start=1):
            ws_header.cell(row=1, column=col).value = header

    if "BOMItem" not in wb.sheetnames:
        ws_item = wb.create_sheet("BOMItem")
        encabezados_item = ["BOMID","POSNR","POSTP","IDNRK","MENGE","MEINS","SORTF","POTX1","POTX2"]
        for col, header in enumerate(encabezados_item, start=1):
            ws_item.cell(row=1, column=col).value = header

    columnas_numericas = ["LEVEL", "ITEM", "QTY","MATERIAL","DESCRIPTION IN CHINESE","SORTSTRNG"]
    mapa_columnas = {
        str(ws.cell(row=1, column=c).value).strip().upper(): openpyxl.utils.get_column_letter(c)
        for c in range(1, ws.max_column + 1)
        if str(ws.cell(row=1, column=c).value).strip().upper() in columnas_numericas
    }

    for nombre, letra in mapa_columnas.items():
        for cell in ws[letra][1:]:
            if cell.value is not None:
                valor = str(cell.value).strip()
                if valor != "":
                    try:
                        cell.value = float(valor.replace(",", ""))
                    except:
                        pass

    for fila in ws.iter_rows():
        for celda in fila:
            celda.alignment = Alignment(horizontal="left")

    colorear_chino(ws)

    modelo_clean = str(texto_modelo).strip().replace(" ", "").replace("\\", "").replace("/", "")
    nuevo_nombre = f"MB-BMM-{modelo_clean}.xlsx"

    ruta_salida_principal = os.path.join(
        os.path.dirname(ruta_salida_principal),
        nuevo_nombre
    )

    wb.save(ruta_salida_principal)
    print(f"[OK] Proceso completo {ruta_salida_principal}")