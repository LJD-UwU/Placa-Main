import os, time, re, pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from glob import glob
from backend.config.sap_config import EXTRAER_ARCHIVO
from backend.utils.txt_to_xlsx import MAINBOARD_2_FILES_FOLDER

# ---------- UTILIDADES ----------
def set_cell(df, row, col, value):
    df.loc[row - 1, col] = value

def limpiar_item(x):
    if pd.isna(x) or str(x).strip() == "":
        return "X"
    return str(x).strip()

def extraer_codigo_pcb(texto):
    if isinstance(texto, str) and ("PCB" in texto or "印制板" in texto):
        m = re.search(r"\.(\d+)(\\|$)", texto)
        if m: return m.group(1)
    return None

# ---------- LIMPIEZA BASE ----------
def limpiar_excel_mainboard_2(ruta_xlsx):
    wb = openpyxl.load_workbook(ruta_xlsx)
    ws = wb.active
    ws.delete_cols(1)
    ws.delete_cols(9, 26)
    ws.delete_rows(1, 9)
    headers = [
        "LEVEL","ITEM","MATERIAL","DESCRIPTION IN CHINESE","DESCRIPTION IN ENGLISH",
        "QTY","UN","LINE 1","LINE 2","SORTSTRNG"
    ]
    ws.insert_cols(1,1)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1,column=col).value = header
    wb.save(ruta_xlsx)

# ---------- PROCESO PRINCIPAL ----------
def procesar_archivo_mainboard_2(ruta_entrada, ruta_salida):
    nombre_archivo = os.path.splitext(os.path.basename(ruta_entrada))[0]
    df = pd.read_excel(ruta_entrada, engine="openpyxl")

    # Inserción inicial
    df = pd.concat([pd.DataFrame(columns=df.columns, index=range(2)), df], ignore_index=True)
    set_cell(df, 1, "ITEM", "X")
    set_cell(df, 1, "MATERIAL", "3TE")
    set_cell(df, 2, "LEVEL", "1")
    set_cell(df, 2, "ITEM", "")
    set_cell(df, 2, "MATERIAL", nombre_archivo)
    set_cell(df, 3, "LEVEL", "1")
    set_cell(df, 3, "ITEM", "X")
    set_cell(df, 3, "MATERIAL", nombre_archivo)

    filas_protegidas = {0,1,2}

    # LEVEL / ITEM
    df["ITEM"] = df["ITEM"].apply(limpiar_item).astype(str)
    df.loc[~df.index.isin(filas_protegidas), "LEVEL"] = 0
    contador_bloque = 10
    for i, val in df["ITEM"].items():
        if i in filas_protegidas: continue
        if val=="X": contador_bloque=10; continue
        df.at[i,"ITEM"]=str(contador_bloque)
        contador_bloque += 10
    nivel_actual=1
    for i in range(len(df)):
        if i in filas_protegidas: continue
        if df.at[i,"ITEM"]=="X": nivel_actual+=1
        else: df.at[i,"LEVEL"]=nivel_actual+1
    for i in range(1,len(df)):
        if i in filas_protegidas: continue
        if df.at[i,"LEVEL"]==0: df.at[i,"LEVEL"]=df.at[i-1,"LEVEL"]

    # Detectar chino
    if "SORTSTRNG" not in df.columns: df["SORTSTRNG"]=None
    for col in ["LINE 1","LINE 2"]:
        mask = df[col].apply(lambda x: any('\u4e00'<=c<='\u9fff' for c in str(x)))
        df.loc[mask,"SORTSTRNG"]=df.loc[mask,col]
        df.loc[mask,col]=None

    # Submateriales
    df["PCB_CODE"]=df["DESCRIPTION IN CHINESE"].apply(extraer_codigo_pcb)
    lista_pcb=df["PCB_CODE"].dropna().tolist()
    if lista_pcb:
        archivos = glob(os.path.join(EXTRAER_ARCHIVO,"*.xls*"))
        archivo_bom = max(archivos,key=os.path.getmtime)
        df_bom = pd.read_excel(archivo_bom, engine="openpyxl")
        df_bom["PCB_clean"]=df_bom["PCB"].astype(str).str.strip()
        if "USE/NO USE" in df_bom.columns:
            df_bom=df_bom[df_bom["USE/NO USE"].astype(str).str.upper()!="NO USE"]
        mask=df_bom["PCB_clean"].apply(lambda x:any(code in x for code in lista_pcb))
        df_filtrado=df_bom.loc[mask,["PCB","Part #","ZH Description","EN Description","QTY","UNIT"]].reset_index(drop=True)
        col_map={"PCB":"ITEM","Part #":"MATERIAL","ZH Description":"DESCRIPTION IN CHINESE",
                 "EN Description":"DESCRIPTION IN ENGLISH","QTY":"QTY","UNIT":"UN"}
        df_sub = pd.DataFrame([{col_map[c]:row[c] for c in df_filtrado.columns if c in col_map} for _,row in df_filtrado.iterrows()])
        df_sub["LEVEL"]=2
        df = pd.concat([df,df_sub],ignore_index=True)

    # Guardar Excel
    df.drop(columns=["PCB_CODE"],errors="ignore",inplace=True)
    df.to_excel(ruta_salida,index=False)
    wb=load_workbook(ruta_salida)
    ws=wb.active

    # Colorear amarillo filas con chino
    amarillo=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
    for idx,tiene in enumerate(df["SORTSTRNG"].notna(),start=2):
        if tiene:
            for col in range(1,10): ws.cell(row=idx,column=col).fill=amarillo
    ws.title="BOMlist"
    wb.save(ruta_salida)
    print(f"[OK] {os.path.basename(ruta_salida)} procesado")

# ---------- EJECUTAR PARA TODA LA CARPETA ----------
CARPETA_ENTRADA = MAINBOARD_2_FILES_FOLDER
CARPETA_SALIDA = MAINBOARD_2_FILES_FOLDER  # puedes cambiar a otra carpeta si quieres
for archivo in os.listdir(CARPETA_ENTRADA):
    if archivo.lower().endswith(".xlsx"):
        ruta_entrada = os.path.join(CARPETA_ENTRADA, archivo)
        ruta_salida = os.path.join(CARPETA_SALIDA, archivo)
        try:
            limpiar_excel_mainboard_2(ruta_entrada)
            procesar_archivo_mainboard_2(ruta_entrada, ruta_salida)
            time.sleep(0.5)
        except Exception as e:
            print(f"[ERROR] {archivo} → {e}")