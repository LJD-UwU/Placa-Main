import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def comparar_y_resaltar_excel(archivo1: str, archivo2: str, salida: str, hoja1: str = None, hoja2: str = None):
    """
    Compara dos Excel celda por celda y genera un nuevo Excel resaltando diferencias.
    
    Parámetros:
    - archivo1, archivo2: rutas de los archivos a comparar
    - salida: ruta del archivo Excel de salida
    - hoja1, hoja2: nombres de las hojas a comparar (opcional)
    
    Las celdas iguales se marcan en verde y las diferentes en rojo.
    """
    # Leer archivos
    df1 = pd.read_excel(archivo1, sheet_name=hoja1)
    df2 = pd.read_excel(archivo2, sheet_name=hoja2)

    # Tomar primera hoja si se devolvió un diccionario
    if isinstance(df1, dict):
        df1 = list(df1.values())[0]
    if isinstance(df2, dict):
        df2 = list(df2.values())[0]

    # Asegurar columnas comunes
    columnas_comunes = df1.columns.intersection(df2.columns)
    df1 = df1[columnas_comunes]
    df2 = df2[columnas_comunes]

    # Igualar filas
    filas_comunes = max(len(df1), len(df2))
    df1 = df1.reindex(range(filas_comunes))
    df2 = df2.reindex(range(filas_comunes))

    # Crear libro de salida
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparacion"

    # Colores
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Escribir encabezados
    for col_idx, col_name in enumerate(columnas_comunes, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Escribir datos y resaltar
    for fila in range(filas_comunes):
        for col_idx, col_name in enumerate(columnas_comunes, start=1):
            valor1 = df1.iloc[fila][col_name]
            valor2 = df2.iloc[fila][col_name]

            # Si ambos son NaN, consideramos iguales
            if pd.isna(valor1) and pd.isna(valor2):
                valor = ""
                ws.cell(row=fila + 2, column=col_idx, value=valor).fill = verde
            elif valor1 == valor2:
                ws.cell(row=fila + 2, column=col_idx, value=valor1).fill = verde
            else:
                # Mostramos ambos valores separados por " | "
                ws.cell(row=fila + 2, column=col_idx, value=f"{valor1} | {valor2}").fill = rojo

    # Guardar archivo
    wb.save(salida)
    print(f"Archivo de comparación generado: {salida}")


# --- Uso ---
archivo1 = r"C:\Users\admin\Documents\Practicante Archivos Main\Automatizacion\MAINBOARD_FILES\11.xlsx"
archivo2 = r"C:\Users\admin\Documents\Practicante Archivos Main\Automatizacion\MAINBOARD_FILES\22.xlsx"
archivo_salida = r"C:\Users\admin\Documents\Practicante Archivos Main\Automatizacion\MAINBOARD_FILES\comparacion.xlsx"

comparar_y_resaltar_excel(archivo1, archivo2, archivo_salida)
